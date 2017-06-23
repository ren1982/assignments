#Code written by Rainier Robles for use in a course project for Database Systems at Freie Universität Berlin
#In the event that code is taken/adapted from an existing source, the sources are attributed in the comments

import openpyxl #used to handle Excel Worksheets
from dateutil import parser #used to parse a timestamp into a tuple
import calendar

def xlref(row, column):
	"""
	Function takes row and column indices and converts it to an Excel cell name.
	Code adapted from https://stackoverflow.com/questions/31420817/convert-excel-row-column-indices-to-alphanumeric-cell-reference-in-python-openpy
	"""
    return openpyxl.utils.get_column_letter(column) + str(row)

def copytonew(oldsheet,newsheet,oldcolumn,newcolumn):
	"""
	Function systematically (row-to-row) copies values from oldcolumn in oldsheet to newcolumn in newsheet.
	"""
	for row in range(1,sheet.max_row+1):
		oldcellname = xlref(row,oldcolumn)
		newcellname = xlref(row,newcolumn)
		newsheet[newcellname] = oldsheet[oldcellname].value

wb = openpyxl.load_workbook('american-election-tweets.xlsx') #opens our Excel Workbook
sheet = wb.get_sheet_by_name('american-election-tweets') #retrieves the specific Sheet we want

wb.create_sheet(index=0, title='cleaned-tweets') #creates a new Sheet with our 'cleaned' data
newsheet = wb.get_sheet_by_name('cleaned-tweets') #saves this Sheet into a variable

usefulcolumns = [1,2,8,9] #list of all columns we need except time
newcolindex = 1 #initializes column numbers for new worksheet and will be incremented accordingly

for i in usefulcolumns: #copies contents of our useful columns to the new sheet
	copytonew(sheet,newsheet,i,newcolindex)
	newcolindex+=1

timeheadercell = xlref(1,newcolindex) #locates which Excel cell will contain the 'time' header
newsheet[timeheadercell] = 'time' #creates a header for 'time' in the new cell

#the following loop takes the timestamp of a Tweet from the old Worksheet, converts it to UNIX time, then writes the UNIX timestamp in the appropriate cell in the new Sheet
#Source used for help with parsing and converting the timestamps: http://avilpage.com/2014/11/python-unix-timestamp-utc-and-their.html
for i in range(2,sheet.max_row+1):
	thetimestamp = sheet.cell(row=i, column=5).value
	parsedts = parser.parse(thetimestamp)
	converted = calendar.timegm(parsedts.utctimetuple())
	cellname = xlref(i,newcolindex)
	newsheet[cellname] = converted

wb.remove_sheet(sheet) #deletes the Workheet with the "unclean" data
wb.save('cleaned-tweets.xlsx') #saves the copied data into a new Workbook, so we don't overwrite the old data in case something goes horribly wrong