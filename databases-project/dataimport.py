#Code written by Rainier Robles for use in a course project for Database Systems at Freie Universität Berlin
#In the event that code is taken/adapted from an existing source, the sources are attributed in the comments

import psycopg2 #used to access our database
import openpyxl #used to handle Excel Worksheets
import re #used for Regular Expressions in Python

allhashtags = [] #initializes (global) list which takes in every single hashtag to make sure no hashtag is repeated

def xlref(row, column):
	"""
	Function takes row and column indices and converts it to an Excel cell name.
	Code adapted from https://stackoverflow.com/questions/31420817/convert-excel-row-column-indices-to-alphanumeric-cell-reference-in-python-openpy
	"""
    return openpyxl.utils.get_column_letter(column) + str(row)

def tweetdata(rownumber):
	"""
	Function takes a rownumber from an Excel Worksheet and sorts the Tweet data into a tuple which will later be inserted into the database.
	The given column indices are based on how they're organized in cleaned-tweets.xlsx.
	"""
	tweet_id = int(rownumber-1)
	tweet_time = sheet.cell(row=rownumber, column=5).value
	tweet_author = sheet.cell(row=rownumber, column=1).value
	tweet_rts = sheet.cell(row=rownumber, column=3).value
	tweet_favs = sheet.cell(row=rownumber, column=4).value
	tweet_text = sheet.cell(row=rownumber, column=2).value
	values = (tweet_id, tweet_time, tweet_author, tweet_rts, tweet_favs, tweet_text)
	return values
	
def hashtaglist(tweet): #returns a list of hashtags in a given tweet, also saves all hashtags into our global list (without repeats)
	"""
	Function takes the tweet text, parses it with the use of regular expressions to retrieve all hashtags in said tweet, then returns a list containing three different lists.
	stripped is the list of all hashtags in the tweet including repeat occurences, and will be used to get data for the enthaelt table.
	single is the list of all unique hashtags in the tweet, basically making a set out of stripped.
	newhashes is the list of all unique hashtags that have not yet been found and used in previous tweets, and are therefore not in our global allhashtags list
	"""
	global allhashtags
	hashtags = re.findall("""(?:(?<=\s\#)|(?<=^\#)) #symbol before hashtag is either whitespace or start of string
		                     (?![\d_]*\s) #hashtag is NOT made up of a string of numbers and underscores followed by whitespace
		                     (?![\d_]*$) #hashtag is NOT made up of a string of numbers and underscores followed by end of string
		                     (?!\w+\#) #hashtag is NOT immediately followed by another hashtag symbol
		                     \w+ #and here's where we finally get our hashtag, which is composed by letters, numbers, and underscores""", tweet, re.X)
	stripped=[] #list of all hashtags including repeats
	single=[] #list of all unique hashtags in a tweet
	newhashes=[] #list of all hashtags currently not in our database
	for hashtag in hashtags:
		b = hashtag.lower() #we want all our hashtags to be in lowercase
		stripped.append(b)
		if b not in single:
			single.append(b)
		if b not in allhashtags:
			allhashtags.append(b)
			newhashes.append(b)
	return [stripped, single, newhashes]

def zipidhashtag(tweet_id, hashtag_list):
	"""
	Function counts frequency of a hashtag within a given hashtag_list, then zips the hashtag and its frequency with the tweet_id.
	A list of triples containing the tweet_id, the hashtag, and the frequency of this hashtag in the tweet is returned.
	"""
	hashtag_count_pairs = []
	id_hashtag_pairs = []
	while len(hashtag_list) != 0:
		indices = [0] #list of where on the list a particular hashtag is found, initialized with 0 since we start with the first entry on the list
		current_hash = hashtag_list[0]
		for i in range(1,len(hashtag_list)): #goes through the rest of the list to find other locations of the hashtag, if any
			if current_hash == hashtag_list[i]:
				indices.append(i)
		hashtag_count_pairs.append([current_hash, len(indices)]) #stores the hashtag and its frequency in a list
		for index in list(reversed(indices)): #deletes all instances of the current_hash(tag) in the list, then goes back to the start of the loop
			del hashtag_list[index]
	for hashtag_count_pair in hashtag_count_pairs: #zips tweet_id with the hashtag and its frequency
		id_hashtag_pairs.append((tweet_id, hashtag_count_pair[0], hashtag_count_pair[1]))
	return id_hashtag_pairs

def ziphashtags(hashtag_list):
	"""
	Function takes a hashtag_list (without duplicates) and takes the crossproduct of the set with itself, excluding reflexive matches to itself.
	A list of lists containing unique hashtag pairs is returned. Also, we take both possible orders of the hashtag (f.e., (hasha, hashb) and (hashb, hasha)) as they will later be used in potential SQL Queries.
	"""
	tagcount = len(hashtag_list)
	hashtagpairs = []
	for i in range(0,tagcount):
		for j in range(0,tagcount):
			if i != j: #creates a list and appends it to our result list, as long as the indices are different
				hashtagpairs.append([hashtag_list[i], hashtag_list[j]])
	return hashtagpairs

conn_string = "host='#####' port='#####' dbname='#####' user='#####' password='#####'" #so we can connect to our database; actual values edited out in order to protect our data
 	
try:
	#here we make a connection to our database
	conn = psycopg2.connect(conn_string)
	cursor = conn.cursor()
except:
	print ("Not connected!")


wb = openpyxl.load_workbook('cleaned-tweets.xlsx') #opens our Excel Workbook
sheet = wb.get_sheet_by_name('cleaned-tweets') #retrieves the specific Sheet we want

#our various Database insertion commands, one for each of our tables
tweet_insertion = """INSERT INTO public.tweet (id,zeitpunkt,autor,retweets,favourites,text) VALUES (%s, %s, %s, %s, %s, %s)"""
hashtag_insertion = """INSERT INTO public.hashtag (text) VALUES (%s)"""
enthaelt_insertion = """INSERT INTO public.enthaelt (id,text,anzahl) VALUES (%s, %s, %s)"""
trittaufmit_insertion = """INSERT INTO public.trittaufmit (texta,textb,id) VALUES (%s, %s, %s)"""

#the for loop iterates through every row and initiates a number of actions
for i in range(2,sheet.max_row+1):
	#print("Processing entry "+str(i-1)) <-- used only to track and potentially debug our code
	values = tweetdata(i) #parses Excel row into a Tuple for use in tweet_insertion
	myhashtaglist = (hashtaglist(values[5])) #element in index 5 of values tuple is our tweet
	#print(myhashtaglist)
	hashwithrepeats = myhashtaglist[0] #list of all hashtags in the tweet with duplicates
	hashunique = myhashtaglist[1] #list of all hashtags in the tweet without duplicates
	newhashtags = myhashtaglist[2] #list of all hashtags in the tweet that have yet to be added to our database
	hashtagpairs = ziphashtags(hashunique) #prepares list of hashtag-hashtag pairs for use in trittaufmit_insertion
	idhashtagpairs = zipidhashtag(values[0], hashwithrepeats) #prepares list of id-hashtag pairs for use in enthaelt_insertion
	cursor.execute(tweet_insertion, values) #performs tweet_insertion
	for eachhashtag in newhashtags: #performs hashtag_insertion
		cursor.execute(hashtag_insertion, (eachhashtag,))
	for idhashtagpair in idhashtagpairs: #performs enthaelt_insertion
		cursor.execute(enthaelt_insertion, idhashtagpair)
	for hashtagpair in hashtagpairs: #performs trittaufmit_insertion
		tobeinserted = (hashtagpair[0], hashtagpair[1], values[0]) #initializes the data set to be inserted
		cursor.execute(trittaufmit_insertion, tobeinserted)

cursor.close()

conn.commit()

conn.close()